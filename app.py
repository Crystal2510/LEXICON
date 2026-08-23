import streamlit as st
import pandas as pd
import sys
import os
import io

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from src.pipeline import ProductPipeline

# ─── PAGE CONFIG ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="CORTEX — Product Intelligence Platform",
    page_icon="⚙️",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─── GLOBAL CSS ───────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

/* ── Reset & Base ── */
*, *::before, *::after { box-sizing: border-box; }

html, body, .stApp {
    background-color: #FFFFFF !important;
    color: #0A0A0A !important;
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
}

[data-testid="stHeader"],
[data-testid="stToolbar"],
[data-testid="stDecoration"] {
    background-color: #FFFFFF !important;
    border-bottom: 1px solid #F1F5F9;
}

/* ── Remove Streamlit padding ── */
.main .block-container {
    padding: 0 2rem 3rem 2rem !important;
    max-width: 1280px !important;
}

/* ── Headings ── */
h1, h2, h3, h4 {
    color: #0A0A0A !important;
    font-family: 'Inter', sans-serif !important;
    font-weight: 700 !important;
    letter-spacing: -0.02em;
}

p, span, label, div {
    font-family: 'Inter', sans-serif !important;
}

/* ── Hero Section ── */
.cortex-hero {
    background: #FFFFFF;
    border-bottom: 1px solid #F1F5F9;
    padding: 3rem 0 2.5rem 0;
    margin-bottom: 0;
}

.cortex-logo {
    font-size: 2.6rem;
    font-weight: 800;
    letter-spacing: -0.04em;
    color: #0A0A0A;
    line-height: 1;
    margin-bottom: 0.4rem;
}

.cortex-logo span {
    color: #2563EB;
}

.cortex-tagline {
    font-size: 1rem;
    color: #64748B;
    font-weight: 400;
    margin-bottom: 0;
    letter-spacing: 0;
}

.cortex-badge {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    background: #EFF6FF;
    color: #2563EB;
    border: 1px solid #BFDBFE;
    border-radius: 100px;
    font-size: 0.72rem;
    font-weight: 600;
    padding: 3px 10px;
    letter-spacing: 0.04em;
    text-transform: uppercase;
    margin-bottom: 1rem;
}

/* ── Feature Cards (Landing) ── */
.feature-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 1rem;
    margin-top: 2rem;
    margin-bottom: 0.5rem;
}

.feature-card {
    background: #F8FAFC;
    border: 1px solid #E2E8F0;
    border-radius: 12px;
    padding: 1.25rem 1rem;
    text-align: center;
    transition: all 0.15s ease;
}

.feature-card:hover {
    border-color: #2563EB;
    background: #EFF6FF;
}

.feature-card .fc-icon {
    font-size: 1.6rem;
    display: block;
    margin-bottom: 0.5rem;
}

.feature-card .fc-title {
    font-size: 0.82rem;
    font-weight: 600;
    color: #0A0A0A;
    display: block;
    margin-bottom: 0.2rem;
}

.feature-card .fc-desc {
    font-size: 0.72rem;
    color: #94A3B8;
    line-height: 1.4;
}

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {
    gap: 0 !important;
    border-bottom: 2px solid #F1F5F9 !important;
    background: #FFFFFF !important;
    padding: 0 !important;
    margin-bottom: 2rem;
}

.stTabs [data-baseweb="tab"] {
    background: transparent !important;
    border: none !important;
    border-bottom: 2px solid transparent !important;
    margin-bottom: -2px !important;
    padding: 1rem 1.5rem !important;
    font-size: 0.88rem !important;
    font-weight: 500 !important;
    color: #64748B !important;
    border-radius: 0 !important;
    transition: all 0.15s ease;
}

.stTabs [aria-selected="true"] {
    color: #2563EB !important;
    border-bottom: 2px solid #2563EB !important;
    font-weight: 600 !important;
    background: transparent !important;
}

.stTabs [data-baseweb="tab"]:hover {
    color: #0A0A0A !important;
    background: #F8FAFC !important;
}

/* ── Section Headers ── */
.section-label {
    font-size: 0.7rem;
    font-weight: 700;
    letter-spacing: 0.1em;
    text-transform: uppercase;
    color: #94A3B8;
    margin-bottom: 0.5rem;
}

.section-title {
    font-size: 1.25rem;
    font-weight: 700;
    color: #0A0A0A;
    margin-bottom: 0.3rem;
    letter-spacing: -0.02em;
}

.section-sub {
    font-size: 0.85rem;
    color: #64748B;
    margin-bottom: 1.5rem;
}

/* ── Upload Zone ── */
.upload-zone {
    border: 2px dashed #CBD5E1;
    border-radius: 16px;
    padding: 2.5rem 2rem;
    text-align: center;
    background: #F8FAFC;
    margin-bottom: 1.5rem;
    transition: all 0.2s ease;
}

.upload-zone:hover {
    border-color: #2563EB;
    background: #EFF6FF;
}

.upload-zone .uz-icon { font-size: 2rem; display: block; margin-bottom: 0.75rem; }
.upload-zone .uz-title { font-size: 0.95rem; font-weight: 600; color: #0A0A0A; }
.upload-zone .uz-sub { font-size: 0.8rem; color: #94A3B8; margin-top: 0.25rem; }

/* ── Stat Cards ── */
.stat-grid {
    display: grid;
    grid-template-columns: repeat(5, 1fr);
    gap: 1rem;
    margin: 1.5rem 0;
}

.stat-card {
    background: #FFFFFF;
    border: 1px solid #E2E8F0;
    border-radius: 12px;
    padding: 1.25rem 1rem;
    position: relative;
    overflow: hidden;
}

.stat-card::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    background: #2563EB;
    border-radius: 12px 12px 0 0;
}

.stat-card.warn::before { background: #F59E0B; }
.stat-card.danger::before { background: #EF4444; }
.stat-card.success::before { background: #10B981; }
.stat-card.neutral::before { background: #94A3B8; }

.stat-card .sc-label {
    font-size: 0.68rem;
    font-weight: 700;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    color: #94A3B8;
    display: block;
    margin-bottom: 0.4rem;
}

.stat-card .sc-value {
    font-size: 1.6rem;
    font-weight: 800;
    color: #0A0A0A;
    letter-spacing: -0.03em;
    line-height: 1;
    display: block;
    margin-bottom: 0.3rem;
}

.stat-card .sc-sub {
    font-size: 0.72rem;
    color: #94A3B8;
}

/* ── Primary Button ── */
button[kind="primary"], .stButton > button[kind="primary"] {
    background: #2563EB !important;
    color: #FFFFFF !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    font-size: 0.9rem !important;
    padding: 0.6rem 2rem !important;
    letter-spacing: -0.01em !important;
    transition: all 0.15s ease !important;
    box-shadow: 0 1px 3px rgba(37,99,235,0.3) !important;
}

button[kind="primary"]:hover {
    background: #1D4ED8 !important;
    box-shadow: 0 4px 12px rgba(37,99,235,0.4) !important;
}

.stDownloadButton > button {
    background: #FFFFFF !important;
    color: #0A0A0A !important;
    border: 1.5px solid #E2E8F0 !important;
    border-radius: 8px !important;
    font-weight: 500 !important;
    font-size: 0.85rem !important;
    transition: all 0.15s ease !important;
}

.stDownloadButton > button:hover {
    border-color: #2563EB !important;
    color: #2563EB !important;
}

/* ── Progress Bar ── */
.stProgress > div > div {
    background: #2563EB !important;
    border-radius: 100px !important;
}
.stProgress > div {
    background: #F1F5F9 !important;
    border-radius: 100px !important;
    height: 6px !important;
}

/* ── Alerts ── */
.stSuccess {
    background: #F0FDF4 !important;
    border: 1px solid #86EFAC !important;
    border-radius: 8px !important;
    color: #166534 !important;
}
.stInfo {
    background: #EFF6FF !important;
    border: 1px solid #BFDBFE !important;
    border-radius: 8px !important;
    color: #1D4ED8 !important;
}
.stError {
    background: #FEF2F2 !important;
    border: 1px solid #FECACA !important;
    border-radius: 8px !important;
    color: #991B1B !important;
}
.stWarning {
    background: #FFFBEB !important;
    border: 1px solid #FDE68A !important;
    border-radius: 8px !important;
    color: #92400E !important;
}

/* ── Confidence Badge ── */
.conf-badge {
    display: inline-flex;
    align-items: center;
    gap: 5px;
    padding: 4px 10px;
    border-radius: 100px;
    font-size: 0.75rem;
    font-weight: 700;
}
.conf-high { background: #D1FAE5; color: #065F46; }
.conf-mid  { background: #FEF3C7; color: #92400E; }
.conf-low  { background: #FEE2E2; color: #991B1B; }

/* ── Before/After Cards ── */
.ba-card {
    background: #FFFFFF;
    border-radius: 12px;
    padding: 1.25rem;
    border: 1px solid #E2E8F0;
    height: 100%;
}

.ba-card-raw {
    border-left: 4px solid #94A3B8;
}

.ba-card-enriched {
    border-left: 4px solid #2563EB;
}

.ba-header {
    display: flex;
    align-items: center;
    gap: 8px;
    margin-bottom: 1rem;
    padding-bottom: 0.75rem;
    border-bottom: 1px solid #F1F5F9;
}

.ba-label-raw {
    background: #F1F5F9;
    color: #475569;
    font-size: 0.7rem;
    font-weight: 700;
    padding: 3px 8px;
    border-radius: 4px;
    letter-spacing: 0.06em;
    text-transform: uppercase;
}

.ba-label-enriched {
    background: #EFF6FF;
    color: #2563EB;
    font-size: 0.7rem;
    font-weight: 700;
    padding: 3px 8px;
    border-radius: 4px;
    letter-spacing: 0.06em;
    text-transform: uppercase;
}

.ba-row-num {
    font-size: 0.8rem;
    color: #94A3B8;
    font-weight: 600;
}

.ba-field {
    display: flex;
    gap: 0.5rem;
    padding: 0.4rem 0;
    border-bottom: 1px solid #F8FAFC;
    align-items: flex-start;
}

.ba-field-label {
    font-size: 0.72rem;
    font-weight: 600;
    color: #94A3B8;
    text-transform: uppercase;
    letter-spacing: 0.04em;
    min-width: 90px;
    padding-top: 2px;
}

.ba-field-value {
    font-size: 0.82rem;
    color: #0A0A0A;
    line-height: 1.4;
    flex: 1;
}

.ba-field-value.placeholder {
    color: #CBD5E1;
    font-style: italic;
}

.ba-mpn {
    font-family: 'Courier New', monospace;
    background: #F1F5F9;
    padding: 2px 6px;
    border-radius: 4px;
    font-size: 0.78rem;
    color: #0A0A0A;
}

.attr-tag {
    display: inline-flex;
    align-items: center;
    gap: 4px;
    background: #F8FAFC;
    border: 1px solid #E2E8F0;
    border-radius: 6px;
    padding: 3px 8px;
    font-size: 0.72rem;
    color: #475569;
    margin: 2px 2px 2px 0;
}

.attr-tag strong { color: #0A0A0A; }

/* ── Quality Metric Row ── */
.qm-row {
    display: flex;
    align-items: center;
    padding: 0.85rem 0;
    border-bottom: 1px solid #F8FAFC;
    gap: 1rem;
}

.qm-icon { font-size: 1rem; min-width: 24px; text-align: center; }
.qm-label { font-size: 0.85rem; color: #334155; flex: 1; font-weight: 500; }
.qm-bar-wrap { flex: 2; background: #F1F5F9; border-radius: 100px; height: 6px; overflow: hidden; }
.qm-bar { height: 6px; border-radius: 100px; }
.qm-bar.good { background: #10B981; }
.qm-bar.mid { background: #F59E0B; }
.qm-bar.bad { background: #EF4444; }
.qm-pct { font-size: 0.85rem; font-weight: 700; min-width: 48px; text-align: right; color: #0A0A0A; }

/* ── Download Cards ── */
.dl-card {
    background: #F8FAFC;
    border: 1.5px solid #E2E8F0;
    border-radius: 16px;
    padding: 1.75rem 1.5rem;
    text-align: center;
    transition: all 0.15s ease;
}

.dl-card:hover { border-color: #2563EB; }
.dl-card .dl-icon { font-size: 2.5rem; display: block; margin-bottom: 0.75rem; }
.dl-card .dl-title { font-size: 1rem; font-weight: 700; color: #0A0A0A; display: block; margin-bottom: 0.3rem; }
.dl-card .dl-sub { font-size: 0.78rem; color: #64748B; line-height: 1.5; margin-bottom: 1.25rem; }

/* ── DataFrames ── */
.stDataFrame {
    border: 1px solid #E2E8F0 !important;
    border-radius: 10px !important;
    overflow: hidden !important;
}

/* ── Divider ── */
.cortex-divider {
    height: 1px;
    background: #F1F5F9;
    margin: 2rem 0;
}

/* ── Pipeline Steps ── */
.pipeline-steps {
    display: flex;
    align-items: center;
    gap: 0;
    margin: 1.5rem 0;
}
.ps-step {
    display: flex;
    align-items: center;
    gap: 8px;
    background: #F8FAFC;
    border: 1px solid #E2E8F0;
    border-radius: 8px;
    padding: 8px 14px;
    font-size: 0.78rem;
    font-weight: 600;
    color: #475569;
}
.ps-step.active {
    background: #EFF6FF;
    border-color: #BFDBFE;
    color: #2563EB;
}
.ps-step.done {
    background: #F0FDF4;
    border-color: #86EFAC;
    color: #059669;
}
.ps-arrow {
    color: #CBD5E1;
    font-size: 0.9rem;
    padding: 0 6px;
}

/* ── Expander overrides ── */
.streamlit-expanderHeader {
    font-size: 0.88rem !important;
    font-weight: 600 !important;
    color: #0A0A0A !important;
    background: #F8FAFC !important;
    border-radius: 8px !important;
}

/* ── Footer ── */
.cortex-footer {
    border-top: 1px solid #F1F5F9;
    padding: 1.5rem 0;
    margin-top: 3rem;
    display: flex;
    justify-content: space-between;
    align-items: center;
}
.cortex-footer-left { font-size: 0.8rem; font-weight: 700; color: #0A0A0A; }
.cortex-footer-left span { color: #2563EB; }
.cortex-footer-right { font-size: 0.75rem; color: #94A3B8; }

/* ── Hide default streamlit branding ── */
#MainMenu { visibility: hidden; }
footer { visibility: hidden; }
</style>
""", unsafe_allow_html=True)


# ─── HELPERS (unchanged logic) ────────────────────────────────────────────────
def to_xlsx_with_colors(df):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Enriched Products"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    red_fill = PatternFill("solid", fgColor="FEE2E2")
    yellow_fill = PatternFill("solid", fgColor="FEF3C7")
    green_fill = PatternFill("solid", fgColor="D1FAE5")

    for col_idx, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        needs_review = str(row.get('NEEDS_REVIEW', 'No'))
        confidence_str = str(row.get('CONFIDENCE_SCORE', '0%'))
        try:
            confidence = int(confidence_str.replace('%', ''))
        except (ValueError, AttributeError):
            confidence = 0

        if needs_review == 'Yes' or confidence < 50:
            row_fill = red_fill
        elif confidence < 80:
            row_fill = yellow_fill
        else:
            row_fill = green_fill

        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
            cell.fill = row_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for col_idx in range(1, len(df.columns) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ''))
            for r in range(1, min(len(df) + 2, 52))
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def init_session_state():
    for key in ['pipeline', 'input_df', 'output_df', 'processing']:
        if key not in st.session_state:
            st.session_state[key] = None if key != 'processing' else False


def render_conf_badge(score_str):
    try:
        score = int(str(score_str).replace('%', ''))
    except (ValueError, AttributeError):
        score = 0
    if score >= 80:
        return f'<span class="conf-badge conf-high">&#x2713; {score}% High</span>'
    elif score >= 50:
        return f'<span class="conf-badge conf-mid">&#x26A0; {score}% Review</span>'
    else:
        return f'<span class="conf-badge conf-low">&#x2717; {score}% Low</span>'


def qm_bar_html(pct_str):
    try:
        pct = float(pct_str.replace('%', ''))
    except Exception:
        pct = 0
    cls = "good" if pct >= 80 else ("mid" if pct >= 50 else "bad")
    return f'<div class="qm-bar-wrap"><div class="qm-bar {cls}" style="width:{pct}%"></div></div>'


def is_placeholder(val):
    bad = {"-- unbranded --", "-- no unilog brand --", "-- no dib brand --", "--", "-", ""}
    return str(val).strip().lower() in bad


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    init_session_state()

    # ── HERO ──────────────────────────────────────────────────────────────────
    st.markdown("""
    <div class="cortex-hero">
        <div class="cortex-badge">&#x26A1; AI-Powered Enrichment</div>
        <div class="cortex-logo">COR<span>TEX</span></div>
        <div class="cortex-tagline">
            Transform raw distributor catalog rows into 252-field, search-ready product content — in seconds.
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── FEATURE CARDS ─────────────────────────────────────────────────────────
    st.markdown("""
    <div class="feature-grid">
        <div class="feature-card">
            <span class="fc-icon">📁</span>
            <span class="fc-title">Upload CSV</span>
            <span class="fc-desc">Drop your raw product catalog — any size, messy data welcome</span>
        </div>
        <div class="feature-card">
            <span class="fc-icon">⚙️</span>
            <span class="fc-title">Enrich & Classify</span>
            <span class="fc-desc">Brand normalization, taxonomy classification, attribute extraction</span>
        </div>
        <div class="feature-card">
            <span class="fc-icon">📊</span>
            <span class="fc-title">Quality Report</span>
            <span class="fc-desc">Field-level compliance scores, confidence analysis, review flags</span>
        </div>
        <div class="feature-card">
            <span class="fc-icon">⬇️</span>
            <span class="fc-title">Export</span>
            <span class="fc-desc">Download enriched CSV or color-coded XLSX — 252 columns ready</span>
        </div>
    </div>
    <div class="cortex-divider"></div>
    """, unsafe_allow_html=True)

    # ── TABS ──────────────────────────────────────────────────────────────────
    tab_upload, tab_preview, tab_quality, tab_download = st.tabs([
        "📁   Upload & Process",
        "🔁   Before / After",
        "📊   Quality Report",
        "⬇️    Download",
    ])

    # ════════════════════════════════════════════════════════════════════════════
    # TAB 1 — UPLOAD & PROCESS
    # ════════════════════════════════════════════════════════════════════════════
    with tab_upload:

        # ── Section header ──
        st.markdown("""
        <div class="section-label">Step 1</div>
        <div class="section-title">Upload Your Product Data</div>
        <div class="section-sub">Upload a CSV file with 6 columns: Mfg_Part_Num, Part_Desc, E1_Brand, Unilog_Brand, DIB_Brand, Part_Manuf</div>
        """, unsafe_allow_html=True)

        # ── File uploader ──
        uploaded_file = st.file_uploader(
            "Drop CSV here",
            type=["csv"],
            label_visibility="collapsed",
        )

        if uploaded_file is not None:
            try:
                input_df = pd.read_csv(uploaded_file, dtype=str, keep_default_na=False)
                st.session_state.input_df = input_df

                # Stat mini-cards
                st.markdown(f"""
                <div style="display:flex; gap:1rem; margin:1rem 0;">
                    <div style="background:#F0FDF4; border:1px solid #86EFAC; border-radius:10px;
                                padding:0.75rem 1.5rem; flex:1; text-align:center;">
                        <div style="font-size:1.5rem; font-weight:800; color:#065F46;">{len(input_df):,}</div>
                        <div style="font-size:0.7rem; font-weight:700; letter-spacing:0.06em;
                                    text-transform:uppercase; color:#6EE7B7;">Rows Loaded</div>
                    </div>
                    <div style="background:#EFF6FF; border:1px solid #BFDBFE; border-radius:10px;
                                padding:0.75rem 1.5rem; flex:1; text-align:center;">
                        <div style="font-size:1.5rem; font-weight:800; color:#1D4ED8;">{len(input_df.columns)}</div>
                        <div style="font-size:0.7rem; font-weight:700; letter-spacing:0.06em;
                                    text-transform:uppercase; color:#93C5FD;">Columns</div>
                    </div>
                    <div style="background:#FFFBEB; border:1px solid #FDE68A; border-radius:10px;
                                padding:0.75rem 1.5rem; flex:2; text-align:center;">
                        <div style="font-size:0.9rem; font-weight:600; color:#92400E; margin-bottom:2px;">
                            {', '.join(input_df.columns.tolist())}
                        </div>
                        <div style="font-size:0.7rem; font-weight:700; letter-spacing:0.06em;
                                    text-transform:uppercase; color:#FCD34D;">Detected Columns</div>
                    </div>
                </div>
                """, unsafe_allow_html=True)

                with st.expander("📋  Preview uploaded data (first 5 rows)", expanded=False):
                    st.dataframe(input_df.head(5), use_container_width=True)

                st.success("✅  File loaded successfully — ready to enrich")

            except Exception as e:
                st.error(f"❌  Error reading file: {str(e)}")

        # ── Divider ──
        st.markdown('<div class="cortex-divider"></div>', unsafe_allow_html=True)

        # ── Enrich section ──
        st.markdown("""
        <div class="section-label">Step 2</div>
        <div class="section-title">Run the Enrichment Pipeline</div>
        <div class="section-sub">CORTEX will normalize brands, classify products, extract attributes and generate all 252 output fields</div>
        """, unsafe_allow_html=True)

        # Pipeline steps visualization
        has_input = st.session_state.input_df is not None
        has_output = st.session_state.output_df is not None
        st.markdown(f"""
        <div class="pipeline-steps" style="flex-wrap:wrap; gap:6px;">
            <div class="ps-step {'done' if has_input else ''}">
                {'✓' if has_input else '1'} &nbsp; Upload CSV
            </div>
            <div class="ps-arrow">→</div>
            <div class="ps-step {'active' if has_input and not has_output else ('done' if has_output else '')}">
                {'✓' if has_output else '2'} &nbsp; Brand Normalize
            </div>
            <div class="ps-arrow">→</div>
            <div class="ps-step {'active' if has_input and not has_output else ('done' if has_output else '')}">
                {'✓' if has_output else '3'} &nbsp; Classify
            </div>
            <div class="ps-arrow">→</div>
            <div class="ps-step {'active' if has_input and not has_output else ('done' if has_output else '')}">
                {'✓' if has_output else '4'} &nbsp; Extract Attributes
            </div>
            <div class="ps-arrow">→</div>
            <div class="ps-step {'active' if has_input and not has_output else ('done' if has_output else '')}">
                {'✓' if has_output else '5'} &nbsp; Generate Descriptions
            </div>
            <div class="ps-arrow">→</div>
            <div class="ps-step {'done' if has_output else ''}">
                {'✓' if has_output else '6'} &nbsp; 252-Column Output
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)

        deep_sourcing = st.toggle(
            "Enable Deep Web Sourcing (Proof of Concept)",
            value=False,
            help="Searches the live web for missing manufacturer specs. Processes all rows with 5 concurrent requests to avoid rate limiting."
        )
        if deep_sourcing:
            st.info("Deep sourcing enabled — all rows with low confidence will be enriched with live web data (Ref URL + specs).")

        col_l, col_m, col_r = st.columns([1, 2, 1])
        with col_m:
            process_button = st.button(
                "⚙️  Enrich with CORTEX",
                use_container_width=True,
                disabled=(st.session_state.input_df is None or st.session_state.processing),
                type="primary",
            )

        if process_button and st.session_state.input_df is not None:
            st.session_state.processing = True
            st.session_state.output_df = None

            progress_bar = st.progress(0)
            status_text = st.empty()

            def update_progress(current, total, status):
                progress_bar.progress(current / total)
                status_text.markdown(
                    f"<div style='font-size:0.82rem;color:#64748B;text-align:center;'>"
                    f"⚙️ &nbsp; {status}</div>",
                    unsafe_allow_html=True
                )

            try:
                pipeline = ProductPipeline()
                status_text.markdown(
                    "<div style='font-size:0.82rem;color:#64748B;text-align:center;'>"
                    "⚙️ &nbsp; Loading reference data...</div>",
                    unsafe_allow_html=True
                )
                input_data = st.session_state.input_df.to_dict('records')
                pipeline.initialize(input_data=input_data)

                output_df = pipeline.process_dataframe(
                    st.session_state.input_df,
                    progress_callback=update_progress,
                    deep_sourcing=deep_sourcing
                )

                st.session_state.output_df = output_df
                st.session_state.processing = False
                progress_bar.progress(1.0)
                status_text.empty()
                st.success(f"✅  Enriched {len(output_df):,} rows — 252 fields populated per product")

            except Exception as e:
                st.session_state.processing = False
                st.error(f"❌  Pipeline error: {str(e)}")

        # ── Results summary ──
        if st.session_state.output_df is not None:
            st.markdown('<div class="cortex-divider"></div>', unsafe_allow_html=True)
            st.markdown("""
            <div class="section-label">Results</div>
            <div class="section-title">Enrichment Summary</div>
            """, unsafe_allow_html=True)

            df = st.session_state.output_df
            total = len(df)
            classified = int((df['Classpath'] != 'General').sum())
            brand_found = int((df['BRAND_NAME'] != '').sum())
            dup_count = int((df.get('IS_DUPLICATE', pd.Series(['False'] * total)) == 'True').sum()) if 'IS_DUPLICATE' in df.columns else 0
            review_count = int((df.get('NEEDS_REVIEW', pd.Series(['No'] * total)) == 'Yes').sum()) if 'NEEDS_REVIEW' in df.columns else 0
            class_pct = classified * 100 // total
            brand_pct = brand_found * 100 // total

            st.markdown(f"""
            <div class="stat-grid">
                <div class="stat-card neutral">
                    <span class="sc-label">Total Products</span>
                    <span class="sc-value">{total:,}</span>
                    <span class="sc-sub">📦 Rows processed</span>
                </div>
                <div class="stat-card {'success' if class_pct >= 80 else 'warn' if class_pct >= 50 else 'danger'}">
                    <span class="sc-label">Classified</span>
                    <span class="sc-value">{classified:,}</span>
                    <span class="sc-sub">📂 {class_pct}% taxonomy matched</span>
                </div>
                <div class="stat-card {'success' if brand_pct >= 80 else 'warn' if brand_pct >= 50 else 'danger'}">
                    <span class="sc-label">Brand Found</span>
                    <span class="sc-value">{brand_found:,}</span>
                    <span class="sc-sub">🏷️ {brand_pct}% normalized</span>
                </div>
                <div class="stat-card {'neutral' if dup_count == 0 else 'warn'}">
                    <span class="sc-label">Duplicates</span>
                    <span class="sc-value">{dup_count}</span>
                    <span class="sc-sub">🔁 MPN duplicates flagged</span>
                </div>
                <div class="stat-card {'success' if review_count == 0 else 'warn' if review_count < total * 0.3 else 'danger'}">
                    <span class="sc-label">Needs Review</span>
                    <span class="sc-value">{review_count}</span>
                    <span class="sc-sub">⚠️ Low-confidence rows</span>
                </div>
            </div>
            """, unsafe_allow_html=True)

            c1, c2 = st.columns(2)
            with c1:
                with st.expander("📂  Classification breakdown", expanded=False):
                    classpath_counts = df['Classpath'].value_counts().head(10)
                    st.bar_chart(classpath_counts)
            with c2:
                with st.expander("🔍  Output preview (first 10 rows)", expanded=False):
                    preview_cols = ['Mfg_Part_Num', 'BRAND_NAME', 'Classpath', 'INVOICE_DESC', 'CONFIDENCE_SCORE']
                    available = [c for c in preview_cols if c in df.columns]
                    st.dataframe(df[available].head(10), use_container_width=True)

    # ════════════════════════════════════════════════════════════════════════════
    # TAB 2 — BEFORE / AFTER
    # ════════════════════════════════════════════════════════════════════════════
    with tab_preview:
        st.markdown("""
        <div class="section-label">Transformation View</div>
        <div class="section-title">Before / After Enrichment</div>
        <div class="section-sub">See exactly how CORTEX transforms each raw product row into structured, search-ready content</div>
        """, unsafe_allow_html=True)

        if st.session_state.output_df is not None and st.session_state.input_df is not None:
            input_df = st.session_state.input_df
            output_df = st.session_state.output_df

            col_a, col_b, _ = st.columns([1, 1, 2])
            with col_a:
                num_samples = st.number_input("Rows to show", 1, 20, 5, key="ba_samples")
            with col_b:
                start_idx = st.number_input("Starting row", 0, max(0, len(input_df) - 1), 0, key="ba_start")

            st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)

            for i in range(int(start_idx), min(int(start_idx + num_samples), len(input_df))):
                row_in = input_df.iloc[i]
                row_out = output_df.iloc[i]
                confidence = row_out.get('CONFIDENCE_SCORE', '0%')

                col_before, col_after = st.columns(2)

                # ── RAW INPUT ──
                with col_before:
                    st.markdown(f"""
                    <div class="ba-card ba-card-raw">
                        <div class="ba-header">
                            <span class="ba-row-num">Row {i + 1}</span>
                            <span class="ba-label-raw">RAW INPUT</span>
                        </div>
                        <div class="ba-field">
                            <span class="ba-field-label">MPN</span>
                            <span class="ba-field-value">
                                <span class="ba-mpn">{row_in.get('Mfg_Part_Num', '') or '—'}</span>
                            </span>
                        </div>
                        <div class="ba-field">
                            <span class="ba-field-label">Part Desc</span>
                            <span class="ba-field-value">{row_in.get('Part_Desc', '') or '—'}</span>
                        </div>
                        <div class="ba-field">
                            <span class="ba-field-label">E1 Brand</span>
                            <span class="ba-field-value {'placeholder' if is_placeholder(row_in.get('E1_Brand','')) else ''}">
                                {row_in.get('E1_Brand', '') or '—'}
                            </span>
                        </div>
                        <div class="ba-field">
                            <span class="ba-field-label">DIB Brand</span>
                            <span class="ba-field-value {'placeholder' if is_placeholder(row_in.get('DIB_Brand','')) else ''}">
                                {row_in.get('DIB_Brand', '') or '—'}
                            </span>
                        </div>
                        <div class="ba-field">
                            <span class="ba-field-label">Manufacturer</span>
                            <span class="ba-field-value">{row_in.get('Part_Manuf', '') or '—'}</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

                # ── ENRICHED OUTPUT ──
                with col_after:
                    # Build attribute tags
                    attr_tags = ""
                    for j in range(1, 21):
                        label = row_out.get(f'ATTRIBUTE_LABEL {j}', '')
                        value = row_out.get(f'ATTRIBUTE_VALUE {j}', '')
                        uom = row_out.get(f'ATTRIBUTE_UOM {j}', '')
                        if label and value:
                            attr_tags += f'<span class="attr-tag"><strong>{label}</strong> = {value} {uom}</span>'
                        if j >= 8:
                            break

                    st.markdown(f"""
                    <div class="ba-card ba-card-enriched">
                        <div class="ba-header">
                            <span class="ba-row-num">Row {i + 1}</span>
                            <span class="ba-label-enriched">ENRICHED OUTPUT</span>
                            <div style="margin-left:auto">{render_conf_badge(confidence)}</div>
                        </div>
                        <div class="ba-field">
                            <span class="ba-field-label">Brand</span>
                            <span class="ba-field-value" style="font-weight:600;color:#2563EB;">
                                {row_out.get('BRAND_NAME', '') or '—'}
                            </span>
                        </div>
                        <div class="ba-field">
                            <span class="ba-field-label">Manufacturer</span>
                            <span class="ba-field-value">{row_out.get('MANUFACTURER_NAME', '') or '—'}</span>
                        </div>
                        <div class="ba-field">
                            <span class="ba-field-label">Classpath</span>
                            <span class="ba-field-value" style="color:#059669;font-weight:500;">
                                {row_out.get('Classpath', '') or '—'}
                            </span>
                        </div>
                        <div class="ba-field">
                            <span class="ba-field-label">Invoice Desc</span>
                            <span class="ba-field-value" style="font-family:monospace;font-size:0.78rem;">
                                {row_out.get('INVOICE_DESC', '') or '—'}
                            </span>
                        </div>
                        <div class="ba-field">
                            <span class="ba-field-label">Mobile Desc</span>
                            <span class="ba-field-value">{row_out.get('MOBILE_DESC', '') or '—'}</span>
                        </div>
                        {'<div class="ba-field"><span class="ba-field-label">Attributes</span><span class="ba-field-value">' + attr_tags + '</span></div>' if attr_tags else ''}
                    </div>
                    """, unsafe_allow_html=True)

                st.markdown("<div style='height:0.75rem'></div>", unsafe_allow_html=True)

        else:
            st.markdown("""
            <div style="text-align:center; padding:4rem 2rem; background:#F8FAFC;
                        border:2px dashed #E2E8F0; border-radius:16px; margin-top:1rem;">
                <div style="font-size:2.5rem; margin-bottom:1rem;">🔁</div>
                <div style="font-size:1rem; font-weight:600; color:#0A0A0A; margin-bottom:0.5rem;">
                    No data processed yet
                </div>
                <div style="font-size:0.85rem; color:#94A3B8;">
                    Upload a CSV file and run the enrichment pipeline to see the Before / After view
                </div>
            </div>
            """, unsafe_allow_html=True)

    # ════════════════════════════════════════════════════════════════════════════
    # TAB 3 — QUALITY REPORT
    # ════════════════════════════════════════════════════════════════════════════
    with tab_quality:
        st.markdown("""
        <div class="section-label">Quality Assurance</div>
        <div class="section-title">Field Compliance Report</div>
        <div class="section-sub">Automated checks against Unilog content guidelines — character limits, casing rules, fill rates</div>
        """, unsafe_allow_html=True)

        if st.session_state.output_df is not None:
            df = st.session_state.output_df
            total = len(df)

            # ── Compliance metrics ──
            invoice_len = df['INVOICE_DESC'].str.len()
            mobile_len = df['MOBILE_DESC'].str.len()

            metrics = [
                ("📄", "Invoice Desc — ALL CAPS (required)",
                 f"{(df['INVOICE_DESC'] == df['INVOICE_DESC'].str.upper()).mean() * 100:.1f}%"),
                ("📏", "Invoice Desc — ≤ 40 characters (limit)",
                 f"{(invoice_len <= 40).mean() * 100:.1f}%"),
                ("📱", "Mobile Desc — 60–80 chars (optimal)",
                 f"{mobile_len.between(60, 80).mean() * 100:.1f}%"),
                ("🏷️", "Brand Name — identified & normalized",
                 f"{(df['BRAND_NAME'] != '').mean() * 100:.1f}%"),
                ("🏭", "Manufacturer Name — identified",
                 f"{(df['MANUFACTURER_NAME'] != '').mean() * 100:.1f}%"),
                ("📂", "Classpath — classified (not General)",
                 f"{(df['Classpath'] != 'General').mean() * 100:.1f}%"),
                ("🔩", "Attributes — ≥ 1 attribute extracted",
                 f"{(df['ATTRIBUTE_VALUE 1'] != '').mean() * 100:.1f}%"),
                ("🔢", "UNSPSC Code — filled",
                 f"{(df['UNSPSC'] != '').mean() * 100:.1f}%"),
            ]

            col_left, col_right = st.columns([3, 2])

            with col_left:
                st.markdown("<div style='font-size:0.85rem;font-weight:600;color:#0A0A0A;margin-bottom:0.75rem;'>Field Compliance Rates</div>", unsafe_allow_html=True)
                html = ""
                for icon, label, pct in metrics:
                    html += f"""
                    <div class="qm-row">
                        <div class="qm-icon">{icon}</div>
                        <div class="qm-label">{label}</div>
                        {qm_bar_html(pct)}
                        <div class="qm-pct">{pct}</div>
                    </div>
                    """
                st.markdown(html, unsafe_allow_html=True)

            with col_right:
                if 'CONFIDENCE_SCORE' in df.columns:
                    try:
                        confidence_vals = df['CONFIDENCE_SCORE'].str.replace('%', '').astype(float)
                        high = int((confidence_vals >= 80).sum())
                        mid = int(((confidence_vals >= 50) & (confidence_vals < 80)).sum())
                        low = int((confidence_vals < 50).sum())
                        avg = confidence_vals.mean()

                        st.markdown(f"""
                        <div style="font-size:0.85rem;font-weight:600;color:#0A0A0A;margin-bottom:0.75rem;">
                            Confidence Distribution
                        </div>
                        <div style="display:flex; flex-direction:column; gap:0.75rem;">
                            <div style="background:#F0FDF4; border:1px solid #86EFAC; border-radius:10px; padding:1rem;">
                                <div style="font-size:1.4rem; font-weight:800; color:#065F46;">{high:,}</div>
                                <div style="font-size:0.72rem; font-weight:700; color:#6EE7B7; text-transform:uppercase; letter-spacing:0.06em;">
                                    ✓ High Confidence (80–100%)
                                </div>
                            </div>
                            <div style="background:#FFFBEB; border:1px solid #FDE68A; border-radius:10px; padding:1rem;">
                                <div style="font-size:1.4rem; font-weight:800; color:#92400E;">{mid:,}</div>
                                <div style="font-size:0.72rem; font-weight:700; color:#FCD34D; text-transform:uppercase; letter-spacing:0.06em;">
                                    ⚠ Review Suggested (50–79%)
                                </div>
                            </div>
                            <div style="background:#FEF2F2; border:1px solid #FECACA; border-radius:10px; padding:1rem;">
                                <div style="font-size:1.4rem; font-weight:800; color:#991B1B;">{low:,}</div>
                                <div style="font-size:0.72rem; font-weight:700; color:#FCA5A5; text-transform:uppercase; letter-spacing:0.06em;">
                                    ✗ Needs Review (0–49%)
                                </div>
                            </div>
                            <div style="background:#EFF6FF; border:1px solid #BFDBFE; border-radius:10px; padding:1rem; text-align:center;">
                                <div style="font-size:1.8rem; font-weight:800; color:#2563EB;">{avg:.1f}%</div>
                                <div style="font-size:0.72rem; font-weight:700; color:#93C5FD; text-transform:uppercase; letter-spacing:0.06em;">
                                    Average Confidence Score
                                </div>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                    except Exception:
                        pass

            # ── Review flagged rows ──
            st.markdown('<div class="cortex-divider"></div>', unsafe_allow_html=True)
            if 'NEEDS_REVIEW' in df.columns:
                review_df = df[df['NEEDS_REVIEW'] == 'Yes']
                count = len(review_df)
                if count > 0:
                    st.markdown(f"""
                    <div style="display:flex; align-items:center; gap:0.75rem; margin-bottom:1rem;">
                        <span style="font-size:1rem; font-weight:700; color:#0A0A0A;">
                            ⚠️  {count} rows flagged for human review
                        </span>
                        <span style="background:#FEF3C7; color:#92400E; font-size:0.72rem; font-weight:700;
                                    padding:3px 10px; border-radius:100px; letter-spacing:0.04em;">
                            LOW CONFIDENCE
                        </span>
                    </div>
                    """, unsafe_allow_html=True)
                    preview = ['Mfg_Part_Num', 'Part_Desc', 'Classpath', 'BRAND_NAME', 'CONFIDENCE_SCORE']
                    available = [c for c in preview if c in review_df.columns]
                    st.dataframe(review_df[available].head(20), use_container_width=True, hide_index=True)
                else:
                    st.success("✅  All rows have high confidence — no manual review needed")

        else:
            st.markdown("""
            <div style="text-align:center; padding:4rem 2rem; background:#F8FAFC;
                        border:2px dashed #E2E8F0; border-radius:16px; margin-top:1rem;">
                <div style="font-size:2.5rem; margin-bottom:1rem;">📊</div>
                <div style="font-size:1rem; font-weight:600; color:#0A0A0A; margin-bottom:0.5rem;">
                    No report available yet
                </div>
                <div style="font-size:0.85rem; color:#94A3B8;">
                    Process your data to generate the quality compliance report
                </div>
            </div>
            """, unsafe_allow_html=True)

    # ════════════════════════════════════════════════════════════════════════════
    # TAB 4 — DOWNLOAD
    # ════════════════════════════════════════════════════════════════════════════
    with tab_download:
        st.markdown("""
        <div class="section-label">Export</div>
        <div class="section-title">Download Enriched Data</div>
        <div class="section-sub">252-column product catalog ready for Unilog import — choose your format</div>
        """, unsafe_allow_html=True)

        if st.session_state.output_df is not None:
            df = st.session_state.output_df

            st.markdown("""
            <div class="dl-card" style="margin-bottom:1rem;">
                <span class="dl-icon">📄</span>
                <span class="dl-title">CSV Format</span>
                <span class="dl-sub">Standard comma-separated file. 252+ columns, all rows included.</span>
            </div>
            """, unsafe_allow_html=True)
            csv_data = df.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="⬇️  Download CSV",
                data=csv_data,
                file_name="cortex_enriched_output.csv",
                mime="text/csv",
                key="download_csv_btn",
            )

            st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)
            st.markdown("""
            <div class="dl-card" style="margin-bottom:1rem;">
                <span class="dl-icon">🎨</span>
                <span class="dl-title">Color-Coded XLSX</span>
                <span class="dl-sub">
                    <span style="color:#065F46;">■</span> Green = High confidence &nbsp;
                    <span style="color:#92400E;">■</span> Yellow = Review &nbsp;
                    <span style="color:#991B1B;">■</span> Red = Needs attention
                </span>
            </div>
            """, unsafe_allow_html=True)
            xlsx_data = to_xlsx_with_colors(df)
            st.download_button(
                label="⬇️  Download XLSX",
                data=xlsx_data,
                file_name="cortex_enriched_output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_xlsx_btn",
            )

            st.markdown('<div class="cortex-divider"></div>', unsafe_allow_html=True)
            st.markdown("""
            <div style="font-size:0.85rem;font-weight:600;color:#0A0A0A;margin-bottom:0.75rem;">
                Output Preview — Key Fields
            </div>
            """, unsafe_allow_html=True)

            preview_cols = ['Mfg_Part_Num', 'BRAND_NAME', 'MANUFACTURER_NAME', 'Classpath',
                            'INVOICE_DESC', 'MOBILE_DESC', 'CONFIDENCE_SCORE', 'NEEDS_REVIEW']
            available = [c for c in preview_cols if c in df.columns]
            st.dataframe(df[available].head(20), use_container_width=True)

            dup_count = int((df.get('IS_DUPLICATE', pd.Series(['False'] * len(df))) == 'True').sum()) if 'IS_DUPLICATE' in df.columns else 0
            if dup_count > 0:
                st.warning(f"⚠️  Found {dup_count} duplicate MPNs — originals kept, duplicates flagged in IS_DUPLICATE column")

        else:
            st.markdown("""
            <div style="text-align:center; padding:4rem 2rem; background:#F8FAFC;
                        border:2px dashed #E2E8F0; border-radius:16px; margin-top:1rem;">
                <div style="font-size:2.5rem; margin-bottom:1rem;">⬇️</div>
                <div style="font-size:1rem; font-weight:600; color:#0A0A0A; margin-bottom:0.5rem;">
                    Nothing to download yet
                </div>
                <div style="font-size:0.85rem; color:#94A3B8;">
                    Upload and process your product data to enable downloads
                </div>
            </div>
            """, unsafe_allow_html=True)

    # ── FOOTER ────────────────────────────────────────────────────────────────
    st.markdown("""
    <div class="cortex-footer">
        <div class="cortex-footer-left">COR<span>TEX</span> &nbsp; Product Intelligence Platform</div>
        <div class="cortex-footer-right">Unilog Hackathon 2026 &nbsp;·&nbsp; AI-Powered Data Enrichment &nbsp;·&nbsp; v1.0</div>
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
