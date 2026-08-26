"""
LEXICON API — Enterprise Product Intelligence Backend
======================================================
FastAPI server with HITL review queue, stats, and approval endpoints.
"""
import io
import os
import sys
import logging
from typing import List, Dict, Any, Optional

import pandas as pd
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from pydantic import BaseModel

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from src.pipeline import ProductPipeline

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="CORTEX API", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

pipeline: Optional[ProductPipeline] = None
enriched_data: Optional[pd.DataFrame] = None


def get_pipeline() -> ProductPipeline:
    global pipeline
    if pipeline is None:
        pipeline = ProductPipeline()
    return pipeline


def get_enriched_data() -> Optional[pd.DataFrame]:
    global enriched_data
    return enriched_data


def set_enriched_data(df: pd.DataFrame):
    global enriched_data
    enriched_data = df


COLUMN_ALIASES = {
    "Mfg_Part_Num": [
        "mpn", "mfg_part_num", "part_number", "part_num", "mfr_part_number",
        "manufacturer_part_number", "mfg part num", "mfg part number",
        "part number", "part no", "model number", "model_num", "catalog_number",
        "item_number", "sku", "product_number", "prod_num", "prod number",
    ],
    "Part_Desc": [
        "description", "part_desc", "part description", "product_description",
        "prod_desc", "product desc", "item_description", "short_description",
        "desc", "product_name", "item_name", "title", "name",
    ],
    "E1_Brand": [
        "brand", "e1_brand", "e1 brand", "brand_name", "brandname",
        "product_brand", "item_brand", "mfg_brand", "mfr_brand",
    ],
    "Unilog_Brand": [
        "unilog_brand", "unilog brand", "unilog", "ub", "ul_brand",
    ],
    "DIB_Brand": [
        "dib_brand", "dib brand", "dib", "db", "dealer_brand",
    ],
    "Part_Manuf": [
        "manufacturer", "part_manuf", "part manufacturer", "mfr", "mfg",
        "vendor", "supplier", "manufacturer_name", "mfg_name", "company",
        "vendor_name", "supplier_name", "mfr_name", "maker",
    ],
}


def map_columns(df: pd.DataFrame) -> pd.DataFrame:
    mapped = {}
    df_cols_lower = {c.lower().strip(): c for c in df.columns}
    for target, aliases in COLUMN_ALIASES.items():
        if target in df.columns:
            mapped[target] = df[target]
            continue
        found = False
        for alias in aliases:
            alias_lower = alias.lower().strip()
            if alias_lower in df_cols_lower:
                mapped[target] = df[df_cols_lower[alias_lower]]
                found = True
                break
            for col_lower, col_orig in df_cols_lower.items():
                if alias_lower in col_lower or col_lower in alias_lower:
                    mapped[target] = df[col_orig]
                    found = True
                    break
            if found:
                break
        if not found:
            for col_lower, col_orig in df_cols_lower.items():
                target_words = target.replace("_", " ").lower().split()
                col_words = col_lower.split()
                if len(set(target_words) & set(col_words)) >= 1:
                    mapped[target] = df[col_orig]
                    found = True
                    break
        if not found:
            mapped[target] = pd.Series([""] * len(df), index=df.index)
    result = pd.DataFrame(mapped, index=df.index)
    extra_cols = [c for c in df.columns if c not in result.columns]
    for c in extra_cols:
        result[c] = df[c]
    return result


@app.get("/api/health")
async def health():
    data_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
    files = {
        "UniCat_Manufacturer_and_Brand_List.xlsx": os.path.exists(os.path.join(data_dir, "UniCat_Manufacturer_and_Brand_List.xlsx")),
        "Unicat_Lov_v1_0_Updated_With_Remarks.xlsx": os.path.exists(os.path.join(data_dir, "Unicat_Lov_v1_0_Updated_With_Remarks.xlsx")),
        "Unilog_Master_UOM_Standards_Abbreviations_and_Terms.xlsx": os.path.exists(os.path.join(data_dir, "Unilog_Master_UOM_Standards_Abbreviations_and_Terms.xlsx")),
        "Decimal_Fraction.xlsx": os.path.exists(os.path.join(data_dir, "Decimal_Fraction.xlsx")),
    }
    loaded = sum(1 for v in files.values() if v)
    return {"status": "ok", "service": "CORTEX API", "reference_files": files, "loaded": loaded, "total": len(files)}


@app.get("/api/product-image/{brand}/{mpn}")
async def get_product_image(brand: str, mpn: str):
    try:
        from PIL import Image, ImageDraw, ImageFont
        img = Image.new('RGB', (400, 300), color=(245, 245, 245))
        draw = ImageDraw.Draw(img)
        try:
            font_large = ImageFont.truetype("arial.ttf", 20)
            font_small = ImageFont.truetype("arial.ttf", 14)
        except Exception:
            font_large = ImageFont.load_default()
            font_small = ImageFont.load_default()
        draw.text((200, 120), brand.replace("_", " "), fill=(51, 51, 51), font=font_large, anchor="mm")
        draw.text((200, 160), mpn.replace("_", " "), fill=(102, 102, 102), font=font_small, anchor="mm")
        draw.text((200, 200), "Product Image", fill=(153, 153, 153), font=font_small, anchor="mm")
        buffer = io.BytesIO()
        img.save(buffer, format="JPEG")
        buffer.seek(0)
        return StreamingResponse(buffer, media_type="image/jpeg")
    except ImportError:
        raise HTTPException(status_code=501, detail="Pillow not installed")


@app.post("/api/preview")
async def preview_csv(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        df = pd.read_csv(io.BytesIO(contents), dtype=str).fillna("")
        preview_rows = df.head(50).to_dict(orient="records")
        return {
            "columns": list(df.columns),
            "total_rows": len(df),
            "preview": preview_rows,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/upload-reference")
async def upload_reference_files(files: List[UploadFile] = File(...)):
    """Upload reference Excel files (UniCat, LOV, UOM, Fraction) to data/ folder."""
    global pipeline
    data_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
    os.makedirs(data_dir, exist_ok=True)
    
    saved = []
    for f in files:
        fname = f.filename or ""
        if not fname.endswith(('.xlsx', '.xls')):
            continue
        contents = await f.read()
        dest = os.path.join(data_dir, fname)
        with open(dest, "wb") as out:
            out.write(contents)
        saved.append(fname)
    
    if saved:
        pipeline = None
    
    return {"status": "ok", "saved": saved, "message": "Reference files loaded. Restart or re-enrich to use."}


@app.post("/api/enrich")
async def enrich_csv(file: UploadFile = File(...), deep_sourcing: bool = True):
    global enriched_data
    try:
        contents = await file.read()
        df = pd.read_csv(io.BytesIO(contents), dtype=str).fillna("")
        logger.info(f"Received {len(df)} rows, {len(df.columns)} columns: {list(df.columns)}")

        df = map_columns(df)
        logger.info(f"After column mapping: {list(df.columns[:6])}")

        pipe = get_pipeline()
        input_data = df.to_dict("records")
        pipe.initialize(input_data=input_data)
        enriched_df = pipe.process_dataframe(df, deep_sourcing=deep_sourcing)

        enriched_data = enriched_df
        logger.info(f"Enriched data stored: {len(enriched_df)} rows")

        stats = compute_stats(enriched_df)
        rows = enriched_df.to_dict(orient="records")
        return {"rows": rows, "stats": stats, "columns": list(enriched_df.columns)}

    except Exception as e:
        logger.error(f"Pipeline error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/stats")
async def get_stats():
    if enriched_data is None:
        raise HTTPException(status_code=404, detail="No data processed yet")
    return compute_stats(enriched_data)


@app.get("/api/review-queue")
async def get_review_queue():
    if enriched_data is None:
        raise HTTPException(status_code=404, detail="No data processed yet")
    review_rows = enriched_data[enriched_data["NEEDS_REVIEW"] == "Yes"].to_dict(orient="records")
    return {"rows": review_rows, "total": len(review_rows)}


@app.get("/api/all-rows")
async def get_all_rows():
    global enriched_data
    if enriched_data is None or len(enriched_data) == 0:
        raise HTTPException(status_code=404, detail="No data processed yet")
    logger.info(f"Returning all rows: {len(enriched_data)} rows")
    key_cols = [
        "Mfg_Part_Num", "Part_Desc", "BRAND_NAME", "MANUFACTURER_NAME",
        "Classpath", "MOBILE_DESC", "INVOICE_DESC", "SHORT_DESC",
        "UNSPSC", "CONFIDENCE_SCORE", "NEEDS_REVIEW",
    ]
    existing = [c for c in key_cols if c in enriched_data.columns]
    df = enriched_data[existing].copy()
    df["_row_index"] = df.index
    conf_vals = df["CONFIDENCE_SCORE"].str.replace("%", "").astype(float)
    df["_conf_sort"] = conf_vals
    df = df.sort_values("_conf_sort")
    df = df.drop(columns=["_conf_sort"])
    rows = df.to_dict(orient="records")
    return {"rows": rows, "total": len(rows)}


class ApproveRequest(BaseModel):
    row_index: int
    edits: Dict[str, str]


@app.post("/api/approve")
async def approve_row(req: ApproveRequest):
    global enriched_data
    if enriched_data is None:
        raise HTTPException(status_code=404, detail="No data processed yet")

    idx = req.row_index
    if idx < 0 or idx >= len(enriched_data):
        raise HTTPException(status_code=400, detail="Invalid row index")

    for col, val in req.edits.items():
        if col in enriched_data.columns:
            enriched_data.at[idx, col] = val

    conf_score = 0
    cp = enriched_data.at[idx, "Classpath"]
    if cp and cp != "General":
        depth = cp.count('>') + 1
        conf_score += min(30, 10 * depth)
    if enriched_data.at[idx, "BRAND_NAME"]:
        known_brands = {'3m', 'dewalt', 'milwaukee', 'makita', 'bosch', 'ridgid', 'ryobi',
            'stanley', 'craftsman', 'diablo', 'freud', 'mirka', 'leviton', 'lutron',
            'philips', 'ge', 'whirlpool', 'frigidaire', 'lg', 'samsung', 'maytag',
            'trex', 'azek', 'timbertech', 'fiberon', 'moen', 'kohler', 'honeywell',
            'southwire', 'satco', 'kichler', 'lithonia', 'cree', 'senco'}
        brand_lower = enriched_data.at[idx, "BRAND_NAME"].lower().strip().rstrip('\u00ae')
        conf_score += 20 if brand_lower in known_brands else 10
    mobile = str(enriched_data.at[idx, "MOBILE_DESC"])
    mob_len = len(mobile)
    if 60 <= mob_len <= 80:
        conf_score += 15
    elif 40 <= mob_len < 60:
        conf_score += 8
    elif mob_len > 80:
        conf_score += 10
    attr_count = sum(1 for i in range(1, 51) if enriched_data.at[idx, f"ATTRIBUTE_LABEL {i}"])
    if attr_count >= 5:
        conf_score += 15
    elif attr_count >= 3:
        conf_score += 10
    elif attr_count >= 1:
        conf_score += 5
    mpn = str(enriched_data.at[idx, "Mfg_Part_Num"])
    if mpn and len(mpn.strip()) > 2:
        conf_score += 5
    conf_score = min(99, conf_score)
    enriched_data.at[idx, "CONFIDENCE_SCORE"] = f"{conf_score}%"
    enriched_data.at[idx, "NEEDS_REVIEW"] = "Yes" if conf_score < 50 else "No"

    return {"status": "ok", "new_confidence": enriched_data.at[idx, "CONFIDENCE_SCORE"]}


class BatchApproveRequest(BaseModel):
    row_indices: List[int]
    edits: Optional[Dict[str, str]] = None


@app.post("/api/approve-batch")
async def approve_batch(req: BatchApproveRequest):
    global enriched_data
    if enriched_data is None:
        raise HTTPException(status_code=404, detail="No data processed yet")

    approved = 0
    for idx in req.row_indices:
        if idx < 0 or idx >= len(enriched_data):
            continue
        if req.edits:
            for col, val in req.edits.items():
                if col in enriched_data.columns:
                    enriched_data.at[idx, col] = val

        conf_score = 0
        cp = enriched_data.at[idx, "Classpath"]
        if cp and cp != "General":
            depth = cp.count('>') + 1
            conf_score += min(30, 10 * depth)
        if enriched_data.at[idx, "BRAND_NAME"]:
            known_brands = {'3m', 'dewalt', 'milwaukee', 'makita', 'bosch', 'ridgid', 'ryobi',
                'stanley', 'craftsman', 'diablo', 'freud', 'mirka', 'leviton', 'lutron',
                'philips', 'ge', 'whirlpool', 'frigidaire', 'lg', 'samsung', 'maytag',
                'trex', 'azek', 'timbertech', 'fiberon', 'moen', 'kohler', 'honeywell',
                'southwire', 'satco', 'kichler', 'lithonia', 'cree', 'senco'}
            brand_lower = enriched_data.at[idx, "BRAND_NAME"].lower().strip().rstrip('\u00ae')
            conf_score += 20 if brand_lower in known_brands else 10
        mobile = str(enriched_data.at[idx, "MOBILE_DESC"])
        mob_len = len(mobile)
        if 60 <= mob_len <= 80:
            conf_score += 15
        elif 40 <= mob_len < 60:
            conf_score += 8
        elif mob_len > 80:
            conf_score += 10
        attr_count = sum(1 for j in range(1, 51) if enriched_data.at[idx, f"ATTRIBUTE_LABEL {j}"])
        if attr_count >= 5:
            conf_score += 15
        elif attr_count >= 3:
            conf_score += 10
        elif attr_count >= 1:
            conf_score += 5
        mpn = str(enriched_data.at[idx, "Mfg_Part_Num"])
        if mpn and len(mpn.strip()) > 2:
            conf_score += 5
        conf_score = min(99, conf_score)
        enriched_data.at[idx, "CONFIDENCE_SCORE"] = f"{conf_score}%"
        enriched_data.at[idx, "NEEDS_REVIEW"] = "Yes" if conf_score < 50 else "No"
        approved += 1

    return {"status": "ok", "approved": approved}


@app.get("/api/review-queue-grouped")
async def get_review_queue_grouped():
    if enriched_data is None:
        raise HTTPException(status_code=404, detail="No data processed yet")

    review_df = enriched_data[enriched_data["NEEDS_REVIEW"] == "Yes"].copy()
    review_df["_row_index"] = review_df.index

    def get_flag_reasons(row):
        reasons = []
        classpath = row.get("Classpath", "")
        brand = row.get("BRAND_NAME", "")
        mobile = row.get("MOBILE_DESC", "")
        invoice = row.get("INVOICE_DESC", "")
        unspsc = row.get("UNSPSC", "")
        conf = row.get("CONFIDENCE_SCORE", "0%")

        if classpath == "General" or not classpath:
            reasons.append("Classpath not determined")
        if not brand:
            reasons.append("Brand not identified")
        if not mobile:
            reasons.append("No mobile description")
        elif len(mobile) < 60 or len(mobile) > 80:
            reasons.append(f"Mobile desc length {len(mobile)} chars (ideal: 60-80)")
        if not invoice:
            reasons.append("No invoice description")
        if not unspsc:
            reasons.append("UNSPSC code missing")
        return reasons

    review_df["_flag_reasons"] = review_df.apply(get_flag_reasons, axis=1)

    dept_col = "Dept" if "Dept" in review_df.columns else "Classpath"
    groups = {}
    for _, row in review_df.iterrows():
        dept = row.get(dept_col, "Unknown") or "Unknown"
        if dept not in groups:
            groups[dept] = []
        groups[dept].append({
            "_row_index": int(row["_row_index"]),
            "Mfg_Part_Num": row.get("Mfg_Part_Num", ""),
            "Part_Desc": row.get("Part_Desc", ""),
            "BRAND_NAME": row.get("BRAND_NAME", ""),
            "MANUFACTURER_NAME": row.get("MANUFACTURER_NAME", ""),
            "Classpath": row.get("Classpath", ""),
            "MOBILE_DESC": row.get("MOBILE_DESC", ""),
            "INVOICE_DESC": row.get("INVOICE_DESC", ""),
            "UNSPSC": row.get("UNSPSC", ""),
            "CONFIDENCE_SCORE": row.get("CONFIDENCE_SCORE", ""),
            "NEEDS_REVIEW": row.get("NEEDS_REVIEW", ""),
            "_flag_reasons": row["_flag_reasons"],
        })

    result = [{"group": k, "items": v, "count": len(v)} for k, v in sorted(groups.items(), key=lambda x: -len(x[1]))]
    return {"groups": result, "total": len(review_df)}


@app.get("/api/summary")
def get_summary():
    if enriched_data is None or len(enriched_data) == 0:
        raise HTTPException(status_code=404, detail="No enriched data")
    df = enriched_data
    conf_vals = pd.to_numeric(df["CONFIDENCE_SCORE"].str.rstrip("%"), errors="coerce")
    return {
        "total": len(df),
        "brand_fill": float((df["BRAND_NAME"] != "").mean() * 100),
        "classpath_fill": float((df["Classpath"] != "").mean() * 100),
        "unspsc_fill": float((df["UNSPSC"] != "").mean() * 100),
        "avg_confidence": float(conf_vals.mean()),
        "needs_review": int((df["NEEDS_REVIEW"] == "Yes").sum()),
        "top_categories": df["Classpath"].value_counts().head(10).to_dict(),
        "top_brands": df["BRAND_NAME"].value_counts().head(10).to_dict(),
    }


@app.post("/api/re-enrich")
def re_enrich_rows(body: dict):
    global enriched_data
    if enriched_data is None:
        raise HTTPException(status_code=404, detail="No enriched data")

    indices = body.get("row_indices", [])
    if not indices:
        raise HTTPException(status_code=400, detail="No row indices provided")

    reprocessed = 0
    pipe = get_pipeline()
    for idx in indices:
        if idx < 0 or idx >= len(enriched_data):
            continue
        row_data = enriched_data.iloc[idx].to_dict()
        try:
            result = pipe.process_row(row_data, deep_sourcing=False, row_index=idx)
            for col, val in result.items():
                if col in enriched_data.columns:
                    enriched_data.at[idx, col] = val
            reprocessed += 1
        except Exception as e:
            logger.warning(f"Re-enrich failed for row {idx}: {e}")

    return {"status": "ok", "reprocessed": reprocessed, "total_requested": len(indices)}


@app.get("/api/classpath-suggestions")
async def get_classpath_suggestions(q: str = ""):
    if enriched_data is None:
        raise HTTPException(status_code=404, detail="No data processed yet")

    classpaths = enriched_data["Classpath"].dropna().unique().tolist()
    classpaths = [c for c in classpaths if c and c != "General"]

    if q:
        q_lower = q.lower()
        classpaths = [c for c in classpaths if q_lower in c.lower()]

    classpaths.sort()
    return {"suggestions": classpaths[:50]}


@app.post("/api/download/review-items")
async def download_review_items():
    if enriched_data is None:
        raise HTTPException(status_code=404, detail="No data processed yet")

    review_df = enriched_data[enriched_data["NEEDS_REVIEW"] == "Yes"].copy()
    if review_df.empty:
        raise HTTPException(status_code=404, detail="No review items found")

    export_cols = ["Mfg_Part_Num", "Part_Desc", "BRAND_NAME", "MANUFACTURER_NAME",
                   "Classpath", "MOBILE_DESC", "INVOICE_DESC", "UNSPSC",
                   "CONFIDENCE_SCORE", "NEEDS_REVIEW"]
    existing_cols = [c for c in export_cols if c in review_df.columns]
    df_export = review_df[existing_cols].copy()
    df_export.insert(0, "_row_index", df_export.index)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Review Items")
        ws = writer.sheets["Review Items"]
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=lexicon_review_items.xlsx"},
    )


@app.post("/api/download/csv")
async def download_csv(rows: Optional[List[Dict[str, Any]]] = None):
    try:
        if rows:
            df = pd.DataFrame(rows)
        elif enriched_data is not None:
            df = enriched_data
        else:
            raise HTTPException(status_code=404, detail="No data to download")
        buffer = io.StringIO()
        df.to_csv(buffer, index=False)
        buffer.seek(0)
        return StreamingResponse(
            io.BytesIO(buffer.getvalue().encode("utf-8")),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=lexicon_enriched_output.csv"},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/download/csv")
async def download_csv_get():
    try:
        if enriched_data is None:
            raise HTTPException(status_code=404, detail="No data to download")
        buffer = io.StringIO()
        enriched_data.to_csv(buffer, index=False)
        buffer.seek(0)
        return StreamingResponse(
            io.BytesIO(buffer.getvalue().encode("utf-8")),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=lexicon_enriched_output.csv"},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/download/xlsx")
async def download_xlsx(rows: Optional[List[Dict[str, Any]]] = None):
    try:
        if rows:
            df = pd.DataFrame(rows)
        elif enriched_data is not None:
            df = enriched_data
        else:
            raise HTTPException(status_code=404, detail="No data to download")
        buffer = io.BytesIO()

        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Enriched Data")
            ws = writer.sheets["Enriched Data"]

            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color="FFFFFF", size=10)
            header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            if "CONFIDENCE_SCORE" in df.columns:
                col_idx = list(df.columns).index("CONFIDENCE_SCORE") + 1
                green = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
                yellow = PatternFill(start_color="FEF7E0", end_color="FEF7E0", fill_type="solid")
                red = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
                for row_idx in range(2, len(df) + 2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    try:
                        val = float(str(cell.value).replace("%", ""))
                        fill = green if val >= 80 else yellow if val >= 50 else red
                        for c in range(1, len(df.columns) + 1):
                            ws.cell(row=row_idx, column=c).fill = fill
                    except Exception:
                        pass

        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=lexicon_enriched_output.xlsx"},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/download/xlsx")
async def download_xlsx_get():
    try:
        if enriched_data is None:
            raise HTTPException(status_code=404, detail="No data to download")
        buffer = io.BytesIO()

        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            enriched_data.to_excel(writer, index=False, sheet_name="Enriched Data")
            ws = writer.sheets["Enriched Data"]

            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color="FFFFFF", size=10)
            header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            ws.auto_filter.ref = ws.dimensions

        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=lexicon_enriched_output.xlsx"},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/download/unilog")
async def download_unilog():
    global enriched_data
    if enriched_data is None:
        raise HTTPException(status_code=404, detail="No data processed yet")
    try:
        pipe = get_pipeline()
        unilog_df = pipe.get_unilog_export(enriched_data)
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            unilog_df.to_excel(writer, index=False, sheet_name="Unilog Export")
            ws = writer.sheets["Unilog Export"]
            from openpyxl.styles import Font, PatternFill
            header_font = Font(bold=True, color="FFFFFF", size=10)
            header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=unilog_252_export.xlsx"},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def compute_stats(df: pd.DataFrame) -> dict:
    total = len(df)
    if total == 0:
        return {"total": 0}

    conf_vals = df["CONFIDENCE_SCORE"].str.replace("%", "").astype(float)

    histogram = []
    for start in range(0, 100, 10):
        end = start + 10
        count = int(((conf_vals >= start) & (conf_vals < end)).sum())
        histogram.append({"range": f"{start}-{end}%", "count": count})

    class_counts = df["Classpath"].value_counts().head(10).to_dict()
    taxonomy = [{"name": k, "value": v} for k, v in class_counts.items()]

    attr_count = 0
    for i in range(1, 21):
        col = f"ATTRIBUTE_VALUE {i}"
        if col in df.columns:
            attr_count += int((df[col] != "").sum())

    total_possible = total * 20
    attr_pct = round(attr_count * 100 / total_possible, 1) if total_possible > 0 else 0

    classified = int((df["Classpath"] != "General").sum())
    brand_found = int((df["BRAND_NAME"] != "").sum())
    inv_filled = int((df["INVOICE_DESC"] != "").sum())
    unspsc_filled = int((df["UNSPSC"] != "").sum())
    dup_count = int((df.get("IS_DUPLICATE", pd.Series(["False"] * total)) == "True").sum())
    review_count = int((df.get("NEEDS_REVIEW", pd.Series(["No"] * total)) == "Yes").sum())

    return {
        "total": total,
        "classified": classified,
        "class_pct": round(classified * 100 / total, 1),
        "brand_found": brand_found,
        "brand_pct": round(brand_found * 100 / total, 1),
        "inv_filled": inv_filled,
        "inv_pct": round(inv_filled * 100 / total, 1),
        "unspsc_filled": unspsc_filled,
        "unspsc_pct": round(unspsc_filled * 100 / total, 1),
        "attr_pct": attr_pct,
        "dup_count": dup_count,
        "review_count": review_count,
        "conf_avg": round(conf_vals.mean(), 1),
        "conf_high": int((conf_vals >= 80).sum()),
        "conf_mid": int(((conf_vals >= 50) & (conf_vals < 80)).sum()),
        "conf_low": int((conf_vals < 50).sum()),
        "histogram": histogram,
        "taxonomy": taxonomy,
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
